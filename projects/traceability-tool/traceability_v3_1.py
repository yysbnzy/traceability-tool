# -*- coding: utf-8 -*-
"""
测试用例与需求溯源工具 v3.1 - 优化UI布局
"""

import re
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


class TraceabilityTool:
    def __init__(self, root):
        self.root = root
        self.root.title("测试用例与需求溯源工具 v3.1")
        self.root.geometry("900x650")
        self.root.minsize(850, 600)

        self.case_file_path = None
        self.req_file_path = None
        self.case_wb = None
        self.req_wb = None

        # 配置样式
        self.setup_styles()
        self.create_ui()

    def setup_styles(self):
        """配置ttk样式"""
        style = ttk.Style()
        style.configure('Title.TLabel', font=('Microsoft YaHei', 11, 'bold'), foreground='#2c3e50')
        style.configure('Group.TLabelframe', font=('Microsoft YaHei', 10, 'bold'))
        style.configure('Group.TLabelframe.Label', font=('Microsoft YaHei', 10, 'bold'), foreground='#34495e')
        style.configure('Hint.TLabel', font=('Microsoft YaHei', 9), foreground='#7f8c8d')

    def create_ui(self):
        # 主容器 - 使用Canvas支持滚动
        main_container = ttk.Frame(self.root, padding="15")
        main_container.pack(fill=tk.BOTH, expand=True)

        # ===== 标题 =====
        title_frame = ttk.Frame(main_container)
        title_frame.pack(fill=tk.X, pady=(0, 15))
        ttk.Label(title_frame, text="📊 测试用例与需求溯源工具", style='Title.TLabel').pack(side=tk.LEFT)
        ttk.Label(title_frame, text="v3.1", style='Hint.TLabel').pack(side=tk.LEFT, padx=(10, 0))

        # ===== 用例文档配置（顶部通栏）=====
        case_doc_frame = ttk.LabelFrame(main_container, text="📄 用例文档配置", style='Group.TLabelframe', padding="10")
        case_doc_frame.pack(fill=tk.X, pady=(0, 10))

        # 文件选择行
        file_row = ttk.Frame(case_doc_frame)
        file_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(file_row, text="文件路径:", width=10).pack(side=tk.LEFT)
        self.case_file_var = tk.StringVar()
        ttk.Entry(file_row, textvariable=self.case_file_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(file_row, text="📂 浏览", command=self.select_case_file, width=8).pack(side=tk.LEFT)

        # 列选择行
        col_row = ttk.Frame(case_doc_frame)
        col_row.pack(fill=tk.X)

        ttk.Label(col_row, text="Sheet页:", width=8).pack(side=tk.LEFT)
        self.case_sheet = ttk.Combobox(col_row, values=[], width=12, state="readonly")
        self.case_sheet.pack(side=tk.LEFT, padx=(0, 20))
        self.case_sheet.set("默认")

        ttk.Label(col_row, text="用例ID列:", width=10).pack(side=tk.LEFT)
        self.case_id_col = ttk.Combobox(col_row, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=8, state="readonly")
        self.case_id_col.pack(side=tk.LEFT, padx=(0, 20))
        self.case_id_col.set("A")

        ttk.Label(col_row, text="需求ID列:", width=10).pack(side=tk.LEFT)
        self.case_req_col = ttk.Combobox(col_row, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=8, state="readonly")
        self.case_req_col.pack(side=tk.LEFT)
        self.case_req_col.set("B")

        # ===== 溯源过滤配置 =====
        filter_frame = ttk.LabelFrame(main_container, text="🔍 溯源过滤配置", style='Group.TLabelframe', padding="10")
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        filter_row = ttk.Frame(filter_frame)
        filter_row.pack(fill=tk.X)

        ttk.Label(filter_row, text="用例ID过滤:", width=12).pack(side=tk.LEFT)
        self.case_filter_prefix = ttk.Entry(filter_row, width=15)
        self.case_filter_prefix.pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(filter_row, text="(只显示此前缀开头的用例ID)", style='Hint.TLabel').pack(side=tk.LEFT)

        filter_row2 = ttk.Frame(filter_frame)
        filter_row2.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(filter_row2, text="需求ID过滤:", width=12).pack(side=tk.LEFT)
        self.req_filter_prefix = ttk.Entry(filter_row2, width=15)
        self.req_filter_prefix.pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(filter_row2, text="(只显示此前缀开头的需求ID)", style='Hint.TLabel').pack(side=tk.LEFT)

        # 存储控件引用（必须在create_concat_config之前初始化）
        self.concat_widgets = {}

        # ===== 中间左右分栏 =====
        middle_frame = ttk.Frame(main_container)
        middle_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        middle_frame.columnconfigure(0, weight=1)
        middle_frame.columnconfigure(1, weight=1)

        # 左侧：用例ID拼接
        left_frame = ttk.LabelFrame(middle_frame, text="🔗 用例ID拼接配置", style='Group.TLabelframe', padding="10")
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        self.create_concat_config(left_frame, "case")

        # 右侧：需求ID拼接
        right_frame = ttk.LabelFrame(middle_frame, text="🔗 需求ID拼接配置", style='Group.TLabelframe', padding="10")
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        self.create_concat_config(right_frame, "req")

        # ===== 需求文档配置（底部）=====
        req_doc_frame = ttk.LabelFrame(main_container, text="📄 需求文档配置（可选）", style='Group.TLabelframe', padding="10")
        req_doc_frame.pack(fill=tk.X, pady=(0, 15))

        req_row = ttk.Frame(req_doc_frame)
        req_row.pack(fill=tk.X)
        ttk.Label(req_row, text="文件路径:", width=10).pack(side=tk.LEFT)
        self.req_file_var = tk.StringVar()
        ttk.Entry(req_row, textvariable=self.req_file_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(req_row, text="📂 浏览", command=self.select_req_file, width=8).pack(side=tk.LEFT)

        ttk.Label(req_row, text="Sheet页:", width=8).pack(side=tk.LEFT, padx=(20, 0))
        self.req_sheet = ttk.Combobox(req_row, values=[], width=12, state="readonly")
        self.req_sheet.pack(side=tk.LEFT, padx=(0, 20))
        self.req_sheet.set("默认")

        ttk.Label(req_row, text="需求ID列:", width=10).pack(side=tk.LEFT, padx=(0, 0))
        self.req_id_col = ttk.Combobox(req_row, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=8, state="readonly")
        self.req_id_col.pack(side=tk.LEFT)
        self.req_id_col.set("A")

        # ===== 底部按钮 =====
        btn_frame = ttk.Frame(main_container)
        btn_frame.pack(fill=tk.X)

        # 左侧：使用说明
        ttk.Button(btn_frame, text="📖 使用说明", command=self.open_manual, width=15).pack(side=tk.LEFT)

        # 右侧：退出 + 开始分析
        ttk.Button(btn_frame, text="❌ 退出", command=self.root.quit, width=10).pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="🚀 开始溯源分析", command=self.start_analysis, width=20).pack(side=tk.RIGHT, padx=5)

    def create_concat_config(self, parent, prefix):
        """创建拼接配置区域"""
        # 前缀配置
        prefix_frame = ttk.Frame(parent)
        prefix_frame.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(prefix_frame, text="前缀来源:", width=10).pack(side=tk.LEFT)
        source_combo = ttk.Combobox(prefix_frame, values=["固定文本", "用例文档列", "需求文档列"],
                                     width=12, state="readonly")
        source_combo.pack(side=tk.LEFT, padx=5)
        source_combo.set("固定文本")

        value_frame = ttk.Frame(prefix_frame)
        value_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        value_entry = ttk.Entry(value_frame, width=15)
        value_entry.pack(fill=tk.X, expand=True)

        self.concat_widgets[f"{prefix}_prefix_source"] = source_combo
        self.concat_widgets[f"{prefix}_prefix_value"] = value_entry
        self.concat_widgets[f"{prefix}_prefix_frame"] = value_frame

        source_combo.bind("<<ComboboxSelected>>",
                         lambda e, p=prefix, t="prefix": self.on_source_change(p, t))

        # 后缀配置
        suffix_frame = ttk.Frame(parent)
        suffix_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(suffix_frame, text="后缀来源:", width=10).pack(side=tk.LEFT)
        source_combo2 = ttk.Combobox(suffix_frame, values=["固定文本", "用例文档列", "需求文档列"],
                                      width=12, state="readonly")
        source_combo2.pack(side=tk.LEFT, padx=5)
        source_combo2.set("固定文本")

        value_frame2 = ttk.Frame(suffix_frame)
        value_frame2.pack(side=tk.LEFT, fill=tk.X, expand=True)

        value_entry2 = ttk.Entry(value_frame2, width=15)
        value_entry2.pack(fill=tk.X, expand=True)

        self.concat_widgets[f"{prefix}_suffix_source"] = source_combo2
        self.concat_widgets[f"{prefix}_suffix_value"] = value_entry2
        self.concat_widgets[f"{prefix}_suffix_frame"] = value_frame2

        source_combo2.bind("<<ComboboxSelected>>",
                          lambda e, p=prefix, t="suffix": self.on_source_change(p, t))

        # 预览说明
        hint_frame = ttk.Frame(parent)
        hint_frame.pack(fill=tk.X, pady=(15, 0))
        ttk.Label(hint_frame, text="💡 示例: 前缀 + 原始ID + 后缀",
                 style='Hint.TLabel', wraplength=350).pack()

    def on_source_change(self, prefix, field_type):
        """当来源类型改变时切换输入控件"""
        key = f"{prefix}_{field_type}"
        source_combo = self.concat_widgets[f"{key}_source"]
        value_frame = self.concat_widgets[f"{key}_frame"]

        source_type = source_combo.get()

        # 清除旧控件
        for widget in value_frame.winfo_children():
            widget.destroy()

        if source_type == "固定文本":
            new_entry = ttk.Entry(value_frame, width=15)
            new_entry.pack(fill=tk.X, expand=True)
            self.concat_widgets[f"{key}_value"] = new_entry
        else:
            col_combo = ttk.Combobox(value_frame, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"),
                                     width=8, state="readonly")
            col_combo.pack(fill=tk.X, expand=True)
            col_combo.set("A")
            self.concat_widgets[f"{key}_value"] = col_combo

    def get_concat_value(self, prefix, field_type, case_row, req_row):
        """获取拼接值"""
        key = f"{prefix}_{field_type}"
        source = self.concat_widgets[f"{key}_source"].get()
        widget = self.concat_widgets[f"{key}_value"]

        if source == "固定文本":
            return widget.get().strip()
        elif source == "用例文档列":
            col_idx = self.col_to_index(widget.get())
            if case_row and len(case_row) > col_idx:
                val = case_row[col_idx]
                return str(val).strip() if val else ""
            return ""
        elif source == "需求文档列":
            if not self.req_wb:
                return None
            col_idx = self.col_to_index(widget.get())
            if req_row and len(req_row) > col_idx:
                val = req_row[col_idx]
                return str(val).strip() if val else ""
            return ""
        return ""

    def select_case_file(self):
        file_path = filedialog.askopenfilename(
            title="选择用例文档",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")]
        )
        if file_path:
            self.case_file_path = file_path
            self.case_file_var.set(file_path)
            try:
                self.case_wb = load_workbook(file_path, read_only=True, data_only=True)
                # 更新Sheet下拉框
                sheet_names = list(self.case_wb.sheetnames)
                self.case_sheet.config(values=sheet_names)
                self.case_sheet.set(sheet_names[0] if sheet_names else "默认")
            except Exception as e:
                messagebox.showerror("错误", f"加载用例文档失败: {str(e)}")

    def select_req_file(self):
        file_path = filedialog.askopenfilename(
            title="选择需求文档",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")]
        )
        if file_path:
            self.req_file_path = file_path
            self.req_file_var.set(file_path)
            try:
                self.req_wb = load_workbook(file_path, read_only=True, data_only=True)
                # 更新Sheet下拉框
                sheet_names = list(self.req_wb.sheetnames)
                self.req_sheet.config(values=sheet_names)
                self.req_sheet.set(sheet_names[0] if sheet_names else "默认")
            except Exception as e:
                messagebox.showerror("错误", f"加载需求文档失败: {str(e)}")

    def download_template(self):
        """下载用例需求溯源输入模板"""
        try:
            output_path = filedialog.asksaveasfilename(
                title="保存模板文件",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile="用例需求溯源输入模板.xlsx"
            )
            if not output_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "用例清单"

            # 表头
            headers = ["用例ID", "关联需求ID"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 示例数据
            example_data = [
                ["SyTC-VA-0001", "SWRD-MCUSecurity-0001"],
                ["SyTC-VA-0002", "SWRD-MCUSecurity-0002, SWRD-MCUSecurity-0003"],
                ["SyTC-VA-0003", ""],
            ]

            for row_idx, row_data in enumerate(example_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 调整列宽
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40

            # 添加说明Sheet
            ws_help = wb.create_sheet("填写说明")
            help_texts = [
                ["用例需求溯源输入模板 - 填写说明"],
                [],
                ["1. 表头说明"],
                ["   用例ID", "测试用例的唯一标识"],
                ["   关联需求ID", "该用例关联的需求ID，多个需求用逗号/分号/换行分隔"],
                [],
                ["2. 填写规则"],
                ["   - 用例ID: 必填，建议格式如 SyTC-VA-0001"],
                ["   - 需求ID: 可选，多个需求用逗号、分号或换行分隔"],
                ["   - 同一用例可分多行填写，工具会自动聚合"],
                [],
                ["3. 分隔符支持"],
                ["   - 逗号 (,)", "顿号 (、)", "分号 (;)", "换行"],
                [],
                ["4. 示例"],
                ["   用例ID: SyTC-VA-0001"],
                ["   需求ID: SWRD-MCUSecurity-0001, SWRD-MCUSecurity-0002"],
            ]

            for row_idx, row_data in enumerate(help_texts, start=1):
                if len(row_data) == 1:
                    ws_help.cell(row=row_idx, column=1, value=row_data[0])
                else:
                    for col_idx, value in enumerate(row_data, start=1):
                        ws_help.cell(row=row_idx, column=col_idx, value=value)

            ws_help.column_dimensions['A'].width = 20
            ws_help.column_dimensions['B'].width = 50

            wb.save(output_path)
            messagebox.showinfo("完成", f"✅ 模板已保存到:\n{output_path}")

        except Exception as e:
            messagebox.showerror("错误", f"生成模板失败: {str(e)}")

    def open_manual(self):
        """弹窗显示简化版使用说明"""
        manual_window = tk.Toplevel(self.root)
        manual_window.title("📖 使用说明")
        manual_window.geometry("550x500")
        
        # 先设置完全透明，避免闪烁
        manual_window.attributes('-alpha', 0.0)
        manual_window.update_idletasks()
        
        # 居中显示
        width = 550
        height = 500
        x = (manual_window.winfo_screenwidth() // 2) - (width // 2)
        y = (manual_window.winfo_screenheight() // 2) - (height // 2)
        manual_window.geometry(f"{width}x{height}+{x}+{y}")
        
        # 设置完成后再显示为不透明
        manual_window.attributes('-alpha', 1.0)
        
        manual_window.transient(self.root)
        manual_window.grab_set()
        
        # 创建滚动文本框
        text_frame = ttk.Frame(manual_window, padding="10")
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Microsoft YaHei", 10), 
                              padx=10, pady=10, bg="#f8f9fa", relief=tk.FLAT)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # 简化的使用说明内容
        manual_text = """【快速上手】

1. 选择用例文档（Excel）
   • 点击"浏览"选择Excel文件
   • 从下拉框选择Sheet页
   • 设置用例ID列和需求ID列

2. 拼接配置（可选）
   • 用例ID/需求ID可添加前缀/后缀
   • 支持固定文本或从Excel列读取

3. 需求文档（可选但建议）
   • 导入独立的需求清单可计算覆盖率
   • 识别未覆盖的孤儿需求

4. 过滤配置（可选）
   • 用例ID过滤：只分析指定前缀的用例
   • 需求ID过滤：只分析指定前缀的需求

5. 开始分析
   • 点击"开始溯源分析"
   • 选择保存位置，生成报告

【输出报告】

• 双向溯源表：用例与需求的双向关联
• 输入源：原始数据的完整副本
• 异常分析：孤儿用例、孤儿需求、被过滤项
• 需求缺失用例：需求文档有但用例未覆盖

【填写规范】

• 多个需求用逗号、分号、顿号、斜杠或换行分隔
• 同一用例可分多行填写，自动聚合
• 表头占第1行，数据从第2行开始

【常见问题】

Q: 结果为空表？
A: 检查列映射是否正确，确认数据从第2行开始

Q: 覆盖率0%？
A: 确认已导入需求文档且列映射正确

Q: 孤儿用例是什么意思？
A: 该用例未关联任何需求（需求ID列为空）
"""
        
        text_widget.insert(tk.END, manual_text)
        text_widget.config(state=tk.DISABLED)
        
        # 关闭按钮
        btn_frame = ttk.Frame(manual_window, padding="10")
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="关闭", command=manual_window.destroy, width=10).pack()

    def split_requirement_ids(self, text):
        if text is None:
            return []
        normalized = str(text)
        # 支持分隔符：顿号、逗号、分号、句号、空格、换行、斜杠
        normalized = re.sub(r'[，；。/\s\n]+', ',', normalized)
        parts = re.split(r'[,;/]+', normalized)
        ids = [p.strip() for p in parts if p.strip()]
        return ids

    def col_to_index(self, col):
        return ord(col) - ord('A')

    def start_analysis(self):
        if not self.case_file_path:
            messagebox.showerror("错误", "请选择用例文档")
            return

        # 检查是否需要需求文档
        sources = [
            self.concat_widgets["case_prefix_source"].get(),
            self.concat_widgets["case_suffix_source"].get(),
            self.concat_widgets["req_prefix_source"].get(),
            self.concat_widgets["req_suffix_source"].get(),
        ]
        if "需求文档列" in sources and not self.req_wb:
            messagebox.showwarning("提示", "您选择了'需求文档列'作为来源，但未上传需求文档。\n请先上传需求文档，或修改来源为其他选项。")
            return

        case_id_idx = self.col_to_index(self.case_id_col.get())
        case_req_idx = self.col_to_index(self.case_req_col.get())
        req_id_idx = self.col_to_index(self.req_id_col.get())

        try:
            result = self.analyze(case_id_idx, case_req_idx, req_id_idx)
            if result is None:
                # 用户取消了保存对话框，不显示完成提示
                return
            messagebox.showinfo("完成", "✅ 溯源分析已完成！")
        except Exception as e:
            messagebox.showerror("错误", f"分析失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def analyze(self, case_id_idx, case_req_idx, req_id_idx):
        case_to_reqs = {}
        case_to_reqs_raw = {}
        req_to_cases = {}
        case_to_original_reqs = {}  # key=拼接用例ID, value=set(原始需求ID，不拼接前缀后缀)

        # 去重统计输入源数据
        unique_case_ids = set()      # 所有出现过的不重复用例ID（含前缀后缀）
        unique_orphan_req_ids = set()  # 只有需求ID（无用例ID）的不重复需求ID（含前缀后缀）

        case_ws = self.case_wb[self.case_sheet.get()] if self.case_sheet.get() and self.case_sheet.get() != "默认" else self.case_wb.active
        max_col = max(case_id_idx, case_req_idx) + 1

        # 预加载需求文档数据
        req_data_cache = {}
        if self.req_wb:
            req_ws = self.req_wb[self.req_sheet.get()] if self.req_sheet.get() and self.req_sheet.get() != "默认" else self.req_wb.active
            for idx, row in enumerate(req_ws.iter_rows(min_row=2, values_only=True), start=2):
                req_data_cache[idx] = row

        for idx, row in enumerate(case_ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
            if len(row) <= max(case_id_idx, case_req_idx):
                continue

            case_id_raw = str(row[case_id_idx]).strip() if row[case_id_idx] else ''
            req_text = row[case_req_idx]
            req_ids = self.split_requirement_ids(req_text)

            if (not case_id_raw or case_id_raw == 'None') and not req_ids:
                continue  # 既无用例ID也无需求ID，跳过

            req_row_data = req_data_cache.get(idx)

            # 动态获取前缀后缀
            case_prefix = self.get_concat_value("case", "prefix", row, req_row_data)
            case_suffix = self.get_concat_value("case", "suffix", row, req_row_data)

            case_id_full = None
            if case_id_raw and case_id_raw != 'None':
                case_id_full = f"{case_prefix}{case_id_raw}{case_suffix}"
                unique_case_ids.add(case_id_full)  # 去重统计用例ID
                if case_id_full not in case_to_reqs:
                    case_to_reqs[case_id_full] = set()
                    case_to_reqs_raw[case_id_raw] = set()
                    case_to_original_reqs[case_id_full] = set()

            for req_id in req_ids:
                req_prefix = self.get_concat_value("req", "prefix", row, req_row_data)
                req_suffix = self.get_concat_value("req", "suffix", row, req_row_data)
                req_id_full = f"{req_prefix}{req_id}{req_suffix}"

                # 只有需求ID（无用例ID）的行，统计为孤儿需求候选
                if not case_id_full:
                    unique_orphan_req_ids.add(req_id_full)

                if req_id_full not in req_to_cases:
                    req_to_cases[req_id_full] = set()
                if case_id_full:
                    req_to_cases[req_id_full].add(case_id_full)
                    case_to_reqs[case_id_full].add(req_id_full)
                    case_to_reqs_raw[case_id_raw].add(req_id_full)
                    case_to_original_reqs[case_id_full].add(req_id)  # 保存原始需求ID（不拼接前缀后缀）

        # 需求文档所有需求
        all_req_ids_raw = set()
        if self.req_wb:
            req_ws = self.req_wb.active
            for row in req_ws.iter_rows(min_row=2, max_col=req_id_idx+1, values_only=True):
                if len(row) <= req_id_idx:
                    continue
                req_id = str(row[req_id_idx]).strip() if row[req_id_idx] else ''
                if req_id and req_id != 'None':
                    all_req_ids_raw.add(req_id)

        missing_case_reqs = all_req_ids_raw - set(req_to_cases.keys())

        # ===== 应用过滤配置（仅影响双向溯源表显示，不影响异常分析） =====
        # 清理过滤条件（去除首尾空格和不可见字符）
        case_filter = self.case_filter_prefix.get().strip()
        req_filter = self.req_filter_prefix.get().strip()
        # 额外清理：去除零宽字符、换行符等
        case_filter = case_filter.replace('\n', '').replace('\r', '').replace('\t', '').replace('\u200b', '')
        req_filter = req_filter.replace('\n', '').replace('\r', '').replace('\t', '').replace('\u200b', '')

        # 保存过滤前的完整数据用于异常分析
        case_to_reqs_raw_full = dict(case_to_reqs_raw)
        req_to_cases_full = dict(req_to_cases)

        # 记录被过滤掉的用例（用于异常分析）
        filtered_out_cases = {}
        filtered_out_cases_original = {}  # key=拼接用例ID, value=set(原始需求ID，用于异常分析显示)

        if case_filter:
            # 过滤用例：只保留符合前缀的用例（仅影响双向溯源表）
            for case_id, reqs in list(case_to_reqs.items()):
                if not case_id.startswith(case_filter):
                    # 用例ID不符合过滤条件，记录为被过滤
                    filtered_out_cases[case_id] = reqs
                    filtered_out_cases_original[case_id] = case_to_original_reqs.get(case_id, set())
            case_to_reqs = {k: v for k, v in case_to_reqs.items() if k.startswith(case_filter)}
            case_to_reqs_raw = {k: v for k, v in case_to_reqs_raw.items() if k.startswith(case_filter)}
            # 同时更新req_to_cases中的关联
            for req_id, cases in list(req_to_cases.items()):
                filtered_cases = {c for c in cases if c.startswith(case_filter)}
                if filtered_cases:
                    req_to_cases[req_id] = filtered_cases
                else:
                    del req_to_cases[req_id]

        if req_filter:
            # 过滤需求：只保留符合前缀的需求（仅影响双向溯源表）
            for case_id, reqs in list(case_to_reqs.items()):
                filtered_reqs = {r for r in reqs if r.startswith(req_filter)}
                if not filtered_reqs:
                    # 所有需求都被过滤，记录被过滤的用例
                    if case_id not in filtered_out_cases:
                        filtered_out_cases[case_id] = reqs
                        filtered_out_cases_original[case_id] = case_to_original_reqs.get(case_id, set())
                else:
                    case_to_reqs[case_id] = filtered_reqs

            req_to_cases = {k: v for k, v in req_to_cases.items() if k.startswith(req_filter)}
            # 同时更新case_to_reqs中的关联
            for case_id, reqs in list(case_to_reqs.items()):
                filtered_reqs = {r for r in reqs if r.startswith(req_filter)}
                if filtered_reqs:
                    case_to_reqs[case_id] = filtered_reqs
                else:
                    del case_to_reqs[case_id]
            # 同步更新case_to_reqs_raw
            for case_id, reqs in list(case_to_reqs_raw.items()):
                filtered_reqs = {r for r in reqs if r.startswith(req_filter)}
                if filtered_reqs:
                    case_to_reqs_raw[case_id] = filtered_reqs
                else:
                    del case_to_reqs_raw[case_id]

        output_path = filedialog.asksaveasfilename(
            title="保存分析结果",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not output_path:
            return None  # 用户取消

        # 提前计算孤儿用例和需求，避免后面重复计算
        orphan_cases = [c for c, r in case_to_reqs_raw_full.items() if not r] if case_to_reqs_raw_full else [c for c, r in case_to_reqs_raw.items() if not r]
        orphan_reqs = [r for r, c in req_to_cases_full.items() if not c] if req_to_cases_full else [r for r, c in req_to_cases.items() if not c]

        # 计算去重后的输入源总数
        total_input_unique = len(unique_case_ids) + len(unique_orphan_req_ids)

        # ===== 数据完整性校验 =====
        # 新校验逻辑：双向溯源表去重用例数 + 异常分析表条目数 = 去重后的输入源总数
        traceability_entries = len([c for c, r in case_to_reqs.items() if r])  # 双向溯源表去重用例数

        # 异常分析表条目数 = 孤儿用例 + 被过滤用例(排除孤儿) + 孤儿需求
        orphan_case_set = set(orphan_cases)
        filtered_non_orphan = [c for c in filtered_out_cases.keys() if c not in orphan_case_set]
        anomaly_entries = len(orphan_cases) + len(filtered_non_orphan) + len(orphan_reqs)

        total_output = traceability_entries + anomaly_entries

        validation_errors = []
        if total_output != total_input_unique:
            validation_errors.append(
                f"数据完整性校验失败：双向溯源表({traceability_entries}) + 异常分析表({anomaly_entries}) = {total_output} ≠ 输入源去重总数({total_input_unique})"
            )

        if validation_errors:
            error_msg = "数据完整性校验失败！\n\n" + "\n".join(validation_errors) + "\n\n请检查数据或联系开发人员。"
            messagebox.showerror("数据校验警告", error_msg)
            # 继续执行导出，但在Excel中记录校验结果

        self.export(output_path, case_to_reqs, case_to_reqs_raw, req_to_cases,
                   all_req_ids_raw, missing_case_reqs, case_to_reqs_raw_full, req_to_cases_full, filtered_out_cases,
                   validation_errors, total_input_unique, traceability_entries, anomaly_entries, filtered_out_cases_original,
                   unique_case_ids, unique_orphan_req_ids)
        
        return output_path

    def export(self, output_path, case_to_reqs, case_to_reqs_raw, req_to_cases,
               all_req_ids_raw, missing_case_reqs, case_to_reqs_raw_full=None, req_to_cases_full=None, filtered_out_cases=None,
               validation_errors=None, total_valid_rows=0, traceability_entries=0, anomaly_entries=0, filtered_out_cases_original=None,
               unique_case_ids=None, unique_orphan_req_ids=None):
        wb = Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # === Sheet1: 双向溯源表-合并版 (A/B + D/E) ===
        ws1 = wb.active
        ws1.title = "双向溯源表-合并版"
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        sorted_cases = sorted([c for c, r in case_to_reqs.items() if r])
        sorted_reqs = sorted([r for r, c in req_to_cases.items() if c])

        # A/B列：用例→需求（合并单元格版）
        ws1['A1'] = '用例ID'
        ws1['B1'] = '关联需求ID'
        for cell in [ws1['A1'], ws1['B1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        current_row = 2
        for case_id in sorted_cases:
            reqs = sorted(case_to_reqs[case_id])
            start_row = current_row
            for req_id in reqs:
                ws1.cell(row=current_row, column=2, value=req_id)
                ws1.cell(row=current_row, column=2).border = thin_border
                ws1.cell(row=current_row, column=2).alignment = left_align
                current_row += 1
            if len(reqs) > 1:
                ws1.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
            ws1.cell(row=start_row, column=1, value=case_id)
            ws1.cell(row=start_row, column=1).border = thin_border
            ws1.cell(row=start_row, column=1).alignment = center_align

        # D/E列：需求→用例（合并单元格版）
        ws1['D1'] = '需求ID'
        ws1['E1'] = '关联用例ID'
        for cell in [ws1['D1'], ws1['E1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        current_row = 2
        for req_id in sorted_reqs:
            cases = sorted(req_to_cases[req_id])
            start_row = current_row
            for case_id in cases:
                ws1.cell(row=current_row, column=5, value=case_id)
                ws1.cell(row=current_row, column=5).border = thin_border
                ws1.cell(row=current_row, column=5).alignment = left_align
                current_row += 1
            if len(cases) > 1:
                ws1.merge_cells(start_row=start_row, start_column=4, end_row=current_row-1, end_column=4)
            ws1.cell(row=start_row, column=4, value=req_id)
            ws1.cell(row=start_row, column=4).border = thin_border
            ws1.cell(row=start_row, column=4).alignment = center_align

        for col in ['A', 'B', 'D', 'E']:
            ws1.column_dimensions[col].width = 40

        # === Sheet2: 双向溯源表-逗号版 (A/B + D/E) ===
        ws_comma = wb.create_sheet("双向溯源表-逗号版")

        # A/B列：用例→需求（逗号分隔版）
        ws_comma['A1'] = '用例ID'
        ws_comma['B1'] = '关联需求ID'
        for cell in [ws_comma['A1'], ws_comma['B1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        for idx, case_id in enumerate(sorted_cases, start=2):
            reqs = sorted(case_to_reqs[case_id])
            ws_comma.cell(row=idx, column=1, value=case_id)
            ws_comma.cell(row=idx, column=2, value=", ".join(reqs))
            for c in [1, 2]:
                ws_comma.cell(row=idx, column=c).border = thin_border
                ws_comma.cell(row=idx, column=c).alignment = left_align

        # D/E列：需求→用例（逗号分隔版）
        ws_comma['D1'] = '需求ID'
        ws_comma['E1'] = '关联用例ID'
        for cell in [ws_comma['D1'], ws_comma['E1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        for idx, req_id in enumerate(sorted_reqs, start=2):
            cases = sorted(req_to_cases[req_id])
            ws_comma.cell(row=idx, column=4, value=req_id)
            ws_comma.cell(row=idx, column=5, value=", ".join(cases))
            for c in [4, 5]:
                ws_comma.cell(row=idx, column=c).border = thin_border
                ws_comma.cell(row=idx, column=c).alignment = left_align

        for col in ['A', 'B', 'D', 'E']:
            ws_comma.column_dimensions[col].width = 40

        # 输入源
        ws2 = wb.create_sheet("输入源")
        ws2['A1'] = '用例ID'
        ws2['B1'] = '需求ID'
        for cell in [ws2['A1'], ws2['B1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        case_ws = self.case_wb[self.case_sheet.get()] if self.case_sheet.get() and self.case_sheet.get() != "默认" else self.case_wb.active
        max_col = max(self.col_to_index(self.case_id_col.get()),
                     self.col_to_index(self.case_req_col.get())) + 1
        for idx, row in enumerate(case_ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
            case_col = self.col_to_index(self.case_id_col.get())
            req_col = self.col_to_index(self.case_req_col.get())

            # 处理用例ID，避免None显示为'None'字符串
            if len(row) > case_col and row[case_col] is not None:
                case_val = str(row[case_col]).strip()
                if case_val == 'None':
                    case_val = ''
            else:
                case_val = ''

            # 处理需求ID，避免None显示为'None'字符串
            if len(row) > req_col and row[req_col] is not None:
                req_val = str(row[req_col]).strip()
                if req_val == 'None':
                    req_val = ''
            else:
                req_val = ''

            ws2.cell(row=idx, column=1, value=case_val)
            ws2.cell(row=idx, column=2, value=req_val)
            ws2.cell(row=idx, column=1).border = thin_border
            ws2.cell(row=idx, column=2).border = thin_border

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 35

        # 异常分析 - 使用完整数据（不受过滤影响）
        ws3 = wb.create_sheet("异常分析")
        ws3['A1'] = '异常类型'
        ws3['B1'] = 'ID'
        ws3['C1'] = '说明'
        for cell in [ws3['A1'], ws3['B1'], ws3['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2

        # 使用完整数据检测异常，不受过滤影响
        orphan_cases = [c for c, r in case_to_reqs_raw_full.items() if not r] if case_to_reqs_raw_full else [c for c, r in case_to_reqs_raw.items() if not r]
        orphan_reqs = [r for r, c in req_to_cases_full.items() if not c] if req_to_cases_full else [r for r, c in req_to_cases.items() if not c]

        # 孤儿用例
        for case_id in orphan_cases:
            ws3.cell(row=row_idx, column=1, value="孤儿用例")
            ws3.cell(row=row_idx, column=2, value=case_id)
            ws3.cell(row=row_idx, column=3, value="该用例未关联任何需求")
            for c in [1, 2, 3]:
                ws3.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

        # 需求被过滤的用例（新增）
        if filtered_out_cases:
            for case_id, reqs in filtered_out_cases.items():
                if case_id not in orphan_cases:  # 避免重复
                    # 使用原始需求ID显示（不拼接前缀后缀）
                    display_reqs = filtered_out_cases_original.get(case_id, reqs) if filtered_out_cases_original else reqs
                    ws3.cell(row=row_idx, column=1, value="需求被过滤")
                    ws3.cell(row=row_idx, column=2, value=case_id)
                    ws3.cell(row=row_idx, column=3, value=f"关联需求不符合过滤条件: {', '.join(display_reqs)}")
                    for c in [1, 2, 3]:
                        ws3.cell(row=row_idx, column=c).border = thin_border
                    row_idx += 1

        # 孤儿需求
        for req_id in orphan_reqs:
            ws3.cell(row=row_idx, column=1, value="孤儿需求")
            ws3.cell(row=row_idx, column=2, value=req_id)
            ws3.cell(row=row_idx, column=3, value="该需求在用例文档中被引用，但无实际关联用例")
            for c in [1, 2, 3]:
                ws3.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 40
        ws3.column_dimensions['C'].width = 40

        # 需求缺失用例
        ws4 = wb.create_sheet("需求缺失用例")
        ws4['A1'] = '需求ID'
        ws4['B1'] = '状态'
        ws4['C1'] = '说明'
        for cell in [ws4['A1'], ws4['B1'], ws4['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for req_id in sorted(missing_case_reqs):
            ws4.cell(row=row_idx, column=1, value=req_id)
            ws4.cell(row=row_idx, column=2, value="未覆盖")
            ws4.cell(row=row_idx, column=3, value="该需求在需求文档中存在，但未在测试用例中体现")
            for c in [1, 2, 3]:
                ws4.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 50

        # 统计汇总
        ws5 = wb.create_sheet("统计汇总")
        ws5['A1'] = '统计项'
        ws5['B1'] = '数量'
        ws5['C1'] = '说明'
        for cell in [ws5['A1'], ws5['B1'], ws5['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # 使用完整数据统计
        total_cases = len(case_to_reqs_raw_full) if case_to_reqs_raw_full else len(case_to_reqs_raw)
        linked_cases = len([c for c, r in (case_to_reqs_raw_full.items() if case_to_reqs_raw_full else case_to_reqs_raw.items()) if r])
        orphan_case_count = len(orphan_cases)
        filtered_case_count = len(filtered_out_cases) if filtered_out_cases else 0
        total_reqs = len(req_to_cases_full) if req_to_cases_full else len(req_to_cases)
        linked_reqs = len([r for r, c in (req_to_cases_full.items() if req_to_cases_full else req_to_cases.items()) if c])
        orphan_req_count = len(orphan_reqs)
        coverage = (linked_reqs / len(all_req_ids_raw) * 100) if all_req_ids_raw else 0

        stats = [
            ('输入源去重用例数', len(unique_case_ids) if unique_case_ids else 0, '用例文档中不重复的用例ID总数'),
            ('输入源去重需求数', len(unique_orphan_req_ids) if unique_orphan_req_ids else 0, '用例文档中不重复的需求ID总数（含孤儿需求）'),
            ('输入源去重总数', total_valid_rows, '用例文档中不重复用例ID + 不重复孤儿需求ID总数'),
            ('', '', ''),
            ('双向溯源表去重用例数', traceability_entries, '过滤后双向溯源表中的用例数量（去重）'),
            ('双向溯源表去重需求数', len([r for r, c in req_to_cases.items() if c]), '过滤后双向溯源表中的需求数量（去重）'),
            ('异常分析表条目数', anomaly_entries, '异常分析表中所有异常条目总数'),
            ('输出数据总计', traceability_entries + anomaly_entries, '双向溯源表 + 异常分析表'),
            ('', '', ''),
            ('总用例数', total_cases, '用例文档中用例ID非空的总数'),
            ('双向关联用例数', linked_cases, '有关联需求的用例数'),
            ('孤儿用例数', orphan_case_count, '未关联任何需求的用例数'),
            ('需求被过滤用例数', filtered_case_count, '关联需求不符合过滤条件的用例数'),
            ('', '', ''),
            ('用例中引用需求数', total_reqs, '用例文档中引用的需求ID总数'),
            ('双向关联需求数', linked_reqs, '有关联用例的需求数'),
            ('孤儿需求数', orphan_req_count, '有用例ID为空但引用的需求数'),
            ('', '', ''),
            ('需求文档总需求数', len(all_req_ids_raw), '需求文档中的需求总数'),
            ('需求缺失用例数', len(missing_case_reqs), '需求文档中有但用例未覆盖的需求数'),
            ('需求覆盖率', f'{coverage:.1f}%', '已覆盖需求数/需求文档总需求数'),
        ]

        row_idx = 2
        for item, count, desc in stats:
            ws5.cell(row=row_idx, column=1, value=item)
            ws5.cell(row=row_idx, column=2, value=count)
            ws5.cell(row=row_idx, column=3, value=desc)
            for c in [1, 2, 3]:
                ws5.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1

        ws5.column_dimensions['A'].width = 25
        ws5.column_dimensions['B'].width = 15
        ws5.column_dimensions['C'].width = 50

        # 数据校验结果显示在统计汇总最后一行
        # 新校验公式：双向溯源表去重用例数 + 异常分析表条目数 = 去重后的输入源总数
        data_integrity_pass = (traceability_entries + anomaly_entries) == total_valid_rows

        # 空行分隔
        row_idx += 1

        # 数据校验结果行
        ws5.cell(row=row_idx, column=1, value='数据完整性校验')
        ws5.cell(row=row_idx, column=2, value='通过' if data_integrity_pass else '失败')
        ws5.cell(row=row_idx, column=3, value=f'双向溯源表去重用例({traceability_entries}) + 异常分析表条目({anomaly_entries}) = {traceability_entries + anomaly_entries}，输入源去重总数={total_valid_rows}')

        # 根据状态设置颜色
        status_cell = ws5.cell(row=row_idx, column=2)
        if data_integrity_pass:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
            status_cell.font = Font(bold=True, color="006100")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
            status_cell.font = Font(bold=True, color="9C0006")

        for c in [1, 2, 3]:
            ws5.cell(row=row_idx, column=c).border = thin_border

        # 如果有校验错误，添加详细说明
        if validation_errors:
            for error in validation_errors:
                row_idx += 1
                ws5.cell(row=row_idx, column=1, value='校验错误')
                ws5.cell(row=row_idx, column=2, value='失败')
                ws5.cell(row=row_idx, column=3, value=error)
                error_cell = ws5.cell(row=row_idx, column=2)
                error_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                error_cell.font = Font(bold=True, color="9C0006")
                for c in [1, 2, 3]:
                    ws5.cell(row=row_idx, column=c).border = thin_border

        wb.save(output_path)
        print(f"溯源分析完成，结果已保存至：{output_path}")


def main():
    root = tk.Tk()
    app = TraceabilityTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
