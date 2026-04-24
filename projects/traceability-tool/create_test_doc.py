# -*- coding: utf-8 -*-
"""
创建测试文档
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "测试数据"

# 表头
ws['A1'] = '用例ID'
ws['B1'] = '需求ID'
ws['C1'] = '说明'

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ['A1', 'B1', 'C1']:
    ws[cell].fill = header_fill
    ws[cell].font = header_font

# 测试数据
test_data = [
    # 正常用例（有用例ID和需求ID）
    ('v7.10#SyTC-Test-0001', 'SWRD-Test-001', '正常用例-单需求'),
    ('v7.10#SyTC-Test-0002', 'SWRD-Test-002,SWRD-Test-003', '正常用例-多需求'),
    ('v7.10#SyTC-Test-0003', 'SWRD-Test-001', '正常用例-共享需求'),
    
    # 孤儿用例（有用例ID，无需求ID）
    ('v7.10#SyTC-Test-0004', '', '孤儿用例-无需求'),
    ('v7.10#SyTC-Test-0005', None, '孤儿用例-需求为空'),
    
    # 孤儿需求（无用例ID，有需求ID）
    ('', 'SWRD-Orphan-001', '孤儿需求-无用例'),
    (None, 'SWRD-Orphan-002', '孤儿需求-用例为空'),
    
    # 边界情况
    ('v7.10#SyTC-Test-0006', 'SWRD-Test-004', '正常用例'),
]

for idx, (case_id, req_id, desc) in enumerate(test_data, start=2):
    ws.cell(row=idx, column=1, value=case_id)
    ws.cell(row=idx, column=2, value=req_id)
    ws.cell(row=idx, column=3, value=desc)

# 调整列宽
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25

# 保存
output_file = r'C:\Users\Administrator\.openclaw\media\temp\测试用例需求文档.xlsx'
wb.save(output_file)
print(f'测试文档已创建: {output_file}')
print(f'包含 {len(test_data)} 条测试数据')
print()
print('测试场景:')
print('  - 正常用例-单需求: v7.10#SyTC-Test-0001 -> SWRD-Test-001')
print('  - 正常用例-多需求: v7.10#SyTC-Test-0002 -> SWRD-Test-002, SWRD-Test-003')
print('  - 正常用例-共享需求: v7.10#SyTC-Test-0003 -> SWRD-Test-001 (共享)')
print('  - 孤儿用例: v7.10#SyTC-Test-0004, v7.10#SyTC-Test-0005')
print('  - 孤儿需求: SWRD-Orphan-001, SWRD-Orphan-002')
