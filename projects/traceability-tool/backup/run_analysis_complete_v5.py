# -*- coding: utf-8 -*-
"""
测试用例与需求溯源分析 - 完整修复版 v5 (包含输入源工作表)
"""

import sys
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from collections import Counter

def analyze_traceability(case_file=None):
    # 如果没有指定文件，自动查找
    if case_file is None or case_file == '':
        case_file = find_excel_file()
    
    if case_file is None:
        print("错误: 未找到Excel文件！")
        print("请将本程序与Excel文件放在同一目录，或将Excel文件拖放到本程序上运行。")
        input("按回车键退出...")
        return
    
    if not os.path.exists(case_file):
        print(f"错误: 文件不存在: {case_file}")
        print("请将本程序与Excel文件放在同一目录，或将Excel文件拖放到本程序上运行。")
        input("按回车键退出...")
        return
    
    print('='*80)
    print('测试用例与需求溯源分析工具 - 完整修复版 v5')
    print('='*80)
    
    # 加载用例文档
    print(f'\n[1] 加载用例文档: {case_file}')
    try:
        case_wb = load_workbook(case_file, data_only=True)
        case_ws = case_wb.active
        print(f'    工作表名称: {case_ws.title}')
    except Exception as e:
        print(f'    加载失败: {e}')
        input("按回车键退出...")
        return
    
    # 分析文档结构
    print('\n[2] 文档结构分析')
    headers = []
    for row in case_ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h else f'Col_{i}' for i, h in enumerate(row)]
    print(f'    列标题: {headers}')
    
    # 查找用例ID和需求ID的列
    print('\n[3] 列识别')
    case_id_col = None
    req_id_col = None
    
    for row in case_ws.iter_rows(min_row=2, max_row=20, values_only=True):
        for col_idx, val in enumerate(row):
            if val:
                val_str = str(val)
                if 'SyTC' in val_str and case_id_col is None:
                    case_id_col = col_idx
                    col_name = headers[col_idx] if col_idx < len(headers) else f'Col_{col_idx}'
                    print(f'    用例ID列 [{col_name}]: 索引 {col_idx}')
                if 'SWRD' in val_str and req_id_col is None:
                    req_id_col = col_idx
                    col_name = headers[req_id_col] if req_id_col < len(headers) else f'Col_{req_id_col}'
                    print(f'    需求ID列 [{col_name}]: 索引 {req_id_col}')
    
    # 开始溯源分析
    print('\n[4] 开始溯源分析...')
    
    case_to_reqs = {}
    req_to_cases = {}
    orphan_cases = []
    all_reqs_in_input = set()
    raw_data = []  # 保存原始数据用于输入源工作表
    
    for idx, row in enumerate(case_ws.iter_rows(min_row=2, values_only=True), start=2):
        max_col = max(case_id_col if case_id_col is not None else 0, 
                      req_id_col if req_id_col is not None else 0)
        if len(row) <= max_col:
            continue
        
        case_id = str(row[case_id_col]).strip() if case_id_col is not None and row[case_id_col] else ''
        req_text = row[req_id_col] if req_id_col is not None else ''
        
        # 保存原始数据
        raw_data.append((case_id, req_text))
        
        # 解析需求ID
        req_ids = []
        if req_text:
            req_str = str(req_text)
            req_parts = re.split(r'[，,；;\n\r\s]+', req_str)
            for part in req_parts:
                part = part.strip()
                if part and 'SWRD' in part:
                    req_ids.append(part)
                    all_reqs_in_input.add(part)
        
        # 只有当用例ID不为空时才记录用例相关数据
        if case_id and case_id != 'None':
            if case_id not in case_to_reqs:
                case_to_reqs[case_id] = set()
            
            if req_ids:
                case_to_reqs[case_id].update(req_ids)
                for req_id in req_ids:
                    if req_id not in req_to_cases:
                        req_to_cases[req_id] = set()
                    req_to_cases[req_id].add(case_id)
            
            if not req_ids:
                orphan_cases.append(case_id)
    
    # 检测孤儿需求
    orphan_reqs = []
    for req_id in all_reqs_in_input:
        if req_id not in req_to_cases or len(req_to_cases[req_id]) == 0:
            orphan_reqs.append(req_id)
    
    # 统计结果
    print('\n' + '='*80)
    print('[5] 溯源分析结果统计')
    print('='*80)
    
    total_cases = len(case_to_reqs)
    total_reqs = len(req_to_cases)
    orphan_count = len(orphan_cases)
    orphan_req_count = len(orphan_reqs)
    cases_with_reqs = sum(1 for reqs in case_to_reqs.values() if reqs)
    
    print(f'\n    总用例数:           {total_cases}')
    print(f'    总需求数:           {total_reqs}')
    print(f'    孤儿用例数:         {orphan_count}')
    print(f'    孤儿需求数:         {orphan_req_count}')
    if total_cases > 0:
        print(f'    已关联需求用例数:   {cases_with_reqs}')
        print(f'    需求关联率:         {cases_with_reqs/total_cases*100:.1f}%')
    
    if orphan_req_count > 0:
        print(f'\n    孤儿需求清单:')
        for req_id in sorted(orphan_reqs):
            print(f'      - {req_id}')
    
    # 导出结果到Excel
    print('\n[6] 导出分析结果...')
    export_results(case_to_reqs, req_to_cases, orphan_cases, orphan_reqs, raw_data, case_file, headers, case_id_col, req_id_col)
    
    print('\n' + '='*80)
    print('分析完成!')
    print('='*80)
    input("按回车键退出...")

def export_results(case_to_reqs, req_to_cases, orphan_cases, orphan_reqs, raw_data, original_file, headers, case_id_col, req_id_col):
    output_file = original_file.replace('.xlsx', '_溯源分析结果_v5.xlsx')
    
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Sheet 1: 双向溯源表
    ws1 = wb.active
    ws1.title = "双向溯源表"
    ws1['A1'] = '用例ID'
    ws1['B1'] = '关联需求ID'
    ws1['D1'] = '需求ID'
    ws1['E1'] = '关联用例ID'
    for cell in [ws1['A1'], ws1['B1'], ws1['D1'], ws1['E1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # 用例->需求
    row_idx = 2
    for case_id in sorted(case_to_reqs.keys()):
        reqs = case_to_reqs[case_id]
        ws1.cell(row=row_idx, column=1, value=case_id)
        ws1.cell(row=row_idx, column=2, value=", ".join(reqs) if reqs else "(孤儿用例)")
        ws1.cell(row=row_idx, column=1).border = thin_border
        ws1.cell(row=row_idx, column=2).border = thin_border
        row_idx += 1
    
    # 需求->用例
    for idx, req_id in enumerate(sorted(req_to_cases.keys()), start=2):
        cases = req_to_cases[req_id]
        ws1.cell(row=idx, column=4, value=req_id)
        ws1.cell(row=idx, column=5, value=", ".join(sorted(cases)) if cases else "(孤儿需求)")
        ws1.cell(row=idx, column=4).border = thin_border
        ws1.cell(row=idx, column=5).border = thin_border
    
    for col in ['A', 'B', 'D', 'E']:
        ws1.column_dimensions[col].width = 45
    
    # Sheet 2: 异常分析
    ws2 = wb.create_sheet("异常分析")
    ws2['A1'] = '异常类型'
    ws2['B1'] = 'ID'
    ws2['C1'] = '说明'
    for cell in [ws2['A1'], ws2['B1'], ws2['C1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    
    # 孤儿需求
    for req_id in orphan_reqs:
        ws2.cell(row=row_idx, column=1, value="孤儿需求")
        ws2.cell(row=row_idx, column=2, value=req_id)
        ws2.cell(row=row_idx, column=3, value="该需求在输入文档中存在，但没有关联任何测试用例")
        for c in [1, 2, 3]:
            ws2.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1
    
    # 孤儿用例
    for case_id in orphan_cases:
        ws2.cell(row=row_idx, column=1, value="孤儿用例")
        ws2.cell(row=row_idx, column=2, value=case_id)
        ws2.cell(row=row_idx, column=3, value="该用例未关联任何需求")
        for c in [1, 2, 3]:
            ws2.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 50
    
    # Sheet 3: 统计汇总
    ws3 = wb.create_sheet("统计汇总")
    ws3['A1'] = '统计项'
    ws3['B1'] = '数值'
    for cell in [ws3['A1'], ws3['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    stats = [
        ('总用例数', len(case_to_reqs)),
        ('总需求数', len(req_to_cases)),
        ('孤儿用例数', len(orphan_cases)),
        ('孤儿需求数', len(orphan_reqs)),
        ('已关联需求用例数', sum(1 for reqs in case_to_reqs.values() if reqs)),
        ('需求关联率', f"{sum(1 for reqs in case_to_reqs.values() if reqs)/len(case_to_reqs)*100:.1f}%"),
    ]
    
    for idx, (item, value) in enumerate(stats, start=2):
        ws3.cell(row=idx, column=1, value=item)
        ws3.cell(row=idx, column=2, value=value)
        ws3.cell(row=idx, column=1).border = thin_border
        ws3.cell(row=idx, column=2).border = thin_border
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    
    # Sheet 4: 输入源 (新增)
    ws4 = wb.create_sheet("输入源")
    
    # 写入表头
    ws4['A1'] = '用例ID'
    ws4['B1'] = '需求ID'
    for cell in [ws4['A1'], ws4['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # 写入原始数据
    for idx, (case_id, req_text) in enumerate(raw_data, start=2):
        ws4.cell(row=idx, column=1, value=case_id if case_id else '(空)')
        ws4.cell(row=idx, column=2, value=req_text if req_text else '(空)')
        ws4.cell(row=idx, column=1).border = thin_border
        ws4.cell(row=idx, column=2).border = thin_border
    
    ws4.column_dimensions['A'].width = 45
    ws4.column_dimensions['B'].width = 50
    
    wb.save(output_file)
    print(f'    结果已保存: {output_file}')
    print(f'    包含工作表: {wb.sheetnames}')

def find_excel_file():
    """自动查找当前目录下的Excel文件"""
    current_dir = os.getcwd()
    excel_files = []
    
    for file in os.listdir(current_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(file)
    
    if len(excel_files) == 0:
        return None
    elif len(excel_files) == 1:
        return os.path.join(current_dir, excel_files[0])
    else:
        for f in excel_files:
            if '用例' in f or '需求' in f or 'case' in f.lower() or 'req' in f.lower():
                return os.path.join(current_dir, f)
        return os.path.join(current_dir, excel_files[0])

if __name__ == "__main__":
    if len(sys.argv) > 1:
        case_file = sys.argv[1]
    else:
        case_file = None
    analyze_traceability(case_file)
