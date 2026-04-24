# -*- coding: utf-8 -*-
"""
检查重复用例ID对应的需求ID
"""

from openpyxl import load_workbook
from collections import defaultdict

input_file = r'C:\Users\Administrator\.openclaw\media\temp\用例需求文档.xlsx'
wb_input = load_workbook(input_file, data_only=True)
ws_input = wb_input.active

# 收集所有用例ID及其对应的需求ID
case_to_reqs = defaultdict(list)

for idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=True), start=2):
    case_id = str(row[4]).strip() if len(row) > 4 and row[4] else ''
    req_text = str(row[1]) if len(row) > 1 and row[1] else ''
    
    if case_id and case_id != 'None':
        case_to_reqs[case_id].append((idx, req_text))

# 找出重复的用例ID
print('重复的用例ID及其对应的需求:')
print('='*80)

duplicates = {k: v for k, v in case_to_reqs.items() if len(v) > 1}

for case_id, rows in sorted(duplicates.items()):
    print(f'\n用例ID: {case_id} (出现 {len(rows)} 次)')
    print('-'*60)
    
    # 检查需求ID是否相同
    req_ids = [r[1] for r in rows]
    unique_reqs = set(req_ids)
    
    for row_idx, req_text in rows:
        display_text = req_text[:60] if req_text else '(无)'
        print(f'  行{row_idx}: {display_text}')
    
    if len(unique_reqs) == 1:
        first_req = list(unique_reqs)[0][:50] if list(unique_reqs)[0] else '(空)'
        print(f'  -> 需求ID相同: {first_req}')
    else:
        print(f'  -> 需求ID不同! 共 {len(unique_reqs)} 种不同需求')
