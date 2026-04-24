import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def split_requirement_ids(text):
    """解析需求ID字符串，支持多种分隔符"""
    if text is None:
        return []
    normalized = str(text)
    normalized = re.sub(r'[，；。\s\n]+', ',', normalized)
    parts = re.split(r'[,;]+', normalized)
    ids = [p.strip() for p in parts if p.strip()]
    return ids


def analyze_traceability(input_path, output_path, case_prefix="", req_prefix=""):
    # 用openpyxl读取Excel
    wb_src = load_workbook(input_path, data_only=True)
    ws_src = wb_src.active

    case_to_reqs = {}
    req_to_cases = {}
    all_req_ids = set()
    invalid_cases = []
    invalid_reqs = []

    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        case_id = str(row[0]).strip() if row[0] else ''
        req_text = row[1]

        if not case_id or case_id == 'None':
            continue

        # 用例前缀校验
        if case_prefix and not case_id.startswith(case_prefix):
            invalid_cases.append(case_id)
            continue

        req_ids = split_requirement_ids(req_text)
        valid_reqs = []

        for req_id in req_ids:
            all_req_ids.add(req_id)
            if req_prefix and not req_id.startswith(req_prefix):
                invalid_reqs.append((case_id, req_id))
            else:
                valid_reqs.append(req_id)

        if case_id not in case_to_reqs:
            case_to_reqs[case_id] = set()
        case_to_reqs[case_id].update(valid_reqs)

        for req_id in valid_reqs:
            if req_id not in req_to_cases:
                req_to_cases[req_id] = set()
            req_to_cases[req_id].add(case_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "溯源分析"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_font = Font(color="9C0006")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 过滤掉孤儿用例
    sorted_cases = sorted([c for c, reqs in case_to_reqs.items() if len(reqs) > 0])
    sorted_reqs = sorted(req_to_cases.keys())

    # === A/B 列：用例→需求（合并单元格版） ===
    ws['A1'] = '用例ID'
    ws['B1'] = '关联需求ID'
    for cell in [ws['A1'], ws['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    current_row = 2
    for case_id in sorted_cases:
        reqs = sorted(case_to_reqs[case_id])
        start_row = current_row
        for req_id in reqs:
            ws.cell(row=current_row, column=2, value=req_id)
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).alignment = left_align
            current_row += 1
        if len(reqs) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
        ws.cell(row=start_row, column=1, value=case_id)
        ws.cell(row=start_row, column=1).border = thin_border
        ws.cell(row=start_row, column=1).alignment = center_align

    # === D/E 列：需求→用例（合并单元格版） ===
    ws['D1'] = '需求ID'
    ws['E1'] = '关联用例ID'
    for cell in [ws['D1'], ws['E1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    current_row = 2
    for req_id in sorted_reqs:
        cases = sorted(req_to_cases[req_id])
        start_row = current_row
        for case_id in cases:
            ws.cell(row=current_row, column=5, value=case_id)
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).alignment = left_align
            current_row += 1
        if len(cases) > 1:
            ws.merge_cells(start_row=start_row, start_column=4, end_row=current_row-1, end_column=4)
        ws.cell(row=start_row, column=4, value=req_id)
        ws.cell(row=start_row, column=4).border = thin_border
        ws.cell(row=start_row, column=4).alignment = center_align

    # === G/I 列：用例→需求（逗号分隔参考版，跳过H列） ===
    ws['G1'] = '用例ID'
    ws['I1'] = '关联需求ID'
    for cell in [ws['G1'], ws['I1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for idx, case_id in enumerate(sorted_cases, start=2):
        reqs = sorted(case_to_reqs[case_id])
        ws.cell(row=idx, column=7, value=case_id)
        ws.cell(row=idx, column=9, value=", ".join(reqs))
        for c in [7, 9]:
            ws.cell(row=idx, column=c).border = thin_border
            ws.cell(row=idx, column=c).alignment = left_align

    # === K/L 列：需求→用例（逗号分隔参考版） ===
    ws['K1'] = '需求ID'
    ws['L1'] = '关联用例ID'
    for cell in [ws['K1'], ws['L1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for idx, req_id in enumerate(sorted_reqs, start=2):
        cases = sorted(req_to_cases[req_id])
        ws.cell(row=idx, column=11, value=req_id)
        ws.cell(row=idx, column=12, value=", ".join(cases))
        for c in [11, 12]:
            ws.cell(row=idx, column=c).border = thin_border
            ws.cell(row=idx, column=c).alignment = left_align

    # 设置列宽
    for col in ['A', 'B', 'D', 'E', 'G', 'I', 'K', 'L']:
        ws.column_dimensions[col].width = 35
    ws.column_dimensions['H'].width = 5

    # 计算孤儿用例和孤儿需求
    orphan_cases = [c for c, reqs in case_to_reqs.items() if len(reqs) == 0]
    orphan_reqs = sorted([r for r in all_req_ids if r not in req_to_cases])

    # === 异常项 Sheet ===
    ws_warn = wb.create_sheet(title="异常项")
    ws_warn.append(["异常类型", "用例ID", "需求ID", "说明"])
    for cell in ws_warn[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_idx = 2
    for case_id in sorted(set(invalid_cases)):
        ws_warn.cell(row=row_idx, column=1, value="用例ID前缀异常")
        ws_warn.cell(row=row_idx, column=2, value=case_id)
        ws_warn.cell(row=row_idx, column=3, value="-")
        ws_warn.cell(row=row_idx, column=4, value="用例ID未以 '%s' 开头" % case_prefix if case_prefix else "-")
        for c in range(1, 5):
            ws_warn.cell(row=row_idx, column=c).fill = warn_fill
            ws_warn.cell(row=row_idx, column=c).font = warn_font
            ws_warn.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

    for case_id, req_id in invalid_reqs:
        ws_warn.cell(row=row_idx, column=1, value="需求ID前缀异常")
        ws_warn.cell(row=row_idx, column=2, value=case_id)
        ws_warn.cell(row=row_idx, column=3, value=req_id)
        ws_warn.cell(row=row_idx, column=4, value="需求ID未以 '%s' 开头" % req_prefix if req_prefix else "-")
        for c in range(1, 5):
            ws_warn.cell(row=row_idx, column=c).fill = warn_fill
            ws_warn.cell(row=row_idx, column=c).font = warn_font
            ws_warn.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

    for case_id in sorted(orphan_cases):
        ws_warn.cell(row=row_idx, column=1, value="孤儿用例")
        ws_warn.cell(row=row_idx, column=2, value=case_id)
        ws_warn.cell(row=row_idx, column=3, value="-")
        ws_warn.cell(row=row_idx, column=4, value="该用例未关联任何有效需求ID")
        for c in range(1, 5):
            ws_warn.cell(row=row_idx, column=c).fill = warn_fill
            ws_warn.cell(row=row_idx, column=c).font = warn_font
            ws_warn.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

    for req_id in orphan_reqs:
        ws_warn.cell(row=row_idx, column=1, value="孤儿需求")
        ws_warn.cell(row=row_idx, column=2, value="-")
        ws_warn.cell(row=row_idx, column=3, value=req_id)
        ws_warn.cell(row=row_idx, column=4, value="该需求ID未被任何用例关联")
        for c in range(1, 5):
            ws_warn.cell(row=row_idx, column=c).fill = warn_fill
            ws_warn.cell(row=row_idx, column=c).font = warn_font
            ws_warn.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

    if row_idx == 2:
        ws_warn.cell(row=row_idx, column=1, value="无异常项")
        ws_warn.cell(row=row_idx, column=4, value="所有数据均通过校验")
        for c in range(1, 5):
            ws_warn.cell(row=row_idx, column=c).border = thin_border

    ws_warn.column_dimensions['A'].width = 22
    ws_warn.column_dimensions['B'].width = 35
    ws_warn.column_dimensions['C'].width = 35
    ws_warn.column_dimensions['D'].width = 40

    wb.save(output_path)

    stats = {
        'case_count': len(case_to_reqs),
        'req_count': len(req_to_cases),
        'orphan_cases': orphan_cases,
        'orphan_reqs': orphan_reqs,
        'invalid_cases': list(set(invalid_cases)),
        'invalid_reqs': invalid_reqs
    }
    return stats


def create_template(output_path):
    """生成输入模板Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "输入模板"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["用例ID", "需求ID"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    examples = [
        ("SyTC-001", "SyRD-001, SyRD-002"),
        ("SyTC-002", "SyRD-002，SyRD-003"),
        ("SyTC-003", "SyRD-001; SyRD-004"),
        ("SyTC-001", "SyRD-005"),
    ]

    for row_idx, (case_id, req_id) in enumerate(examples, start=2):
        ws.cell(row=row_idx, column=1, value=case_id)
        ws.cell(row=row_idx, column=2, value=req_id)
        for c in [1, 2]:
            ws.cell(row=row_idx, column=c).border = thin_border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    ws['D1'] = "填写说明："
    ws['D1'].font = Font(bold=True, size=11)
    notes = [
        "1. 第一列填写用例ID，第二列填写需求ID",
        "2. 一个用例对应多个需求时，可用逗号、分号或中文逗号分隔",
        "3. 同一个用例ID可分多行填写，工具会自动聚合",
        "4. 建议用例ID以SyTC开头，需求ID以SyRD开头",
    ]
    for idx, note in enumerate(notes, start=2):
        ws.cell(row=idx, column=4, value=note)

    ws.column_dimensions['D'].width = 55
    wb.save(output_path)


class TraceabilityGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("用例-需求溯源分析工具")
        self.root.geometry("700x520")
        self.root.resizable(False, False)

        style = ttk.Style()
        style.configure("TButton", font=("Microsoft YaHei", 10))
        style.configure("TLabel", font=("Microsoft YaHei", 10))

        frame_file = ttk.Frame(root, padding=15)
        frame_file.pack(fill=tk.X)

        ttk.Label(frame_file, text="输入Excel文件：").grid(row=0, column=0, sticky=tk.W)
        self.entry_path = ttk.Entry(frame_file, width=50)
        self.entry_path.grid(row=0, column=1, padx=5)
        ttk.Button(frame_file, text="选择文件", command=self.select_file).grid(row=0, column=2)

        frame_prefix = ttk.LabelFrame(root, text="前缀过滤（可选）", padding=10)
        frame_prefix.pack(fill=tk.X, padx=15, pady=5)

        ttk.Label(frame_prefix, text="用例ID前缀：").grid(row=0, column=0, sticky=tk.W)
        self.entry_case_prefix = ttk.Entry(frame_prefix, width=20)
        self.entry_case_prefix.grid(row=0, column=1, padx=5, sticky=tk.W)

        ttk.Label(frame_prefix, text="需求ID前缀：").grid(row=0, column=2, sticky=tk.W, padx=(20, 0))
        self.entry_req_prefix = ttk.Entry(frame_prefix, width=20)
        self.entry_req_prefix.grid(row=0, column=3, padx=5, sticky=tk.W)

        ttk.Label(frame_prefix, text="提示：留空表示不启用前缀过滤", foreground="gray").grid(row=1, column=0, columnspan=4, sticky=tk.W, pady=(5, 0))

        frame_btn = ttk.Frame(root, padding=10)
        frame_btn.pack(fill=tk.X)

        ttk.Button(frame_btn, text="下载模板", command=self.download_template).pack(side=tk.LEFT, padx=(0, 10))
        self.btn_run = ttk.Button(frame_btn, text="开始分析", command=self.run_analysis)
        self.btn_run.pack(side=tk.LEFT)

        frame_log = ttk.LabelFrame(root, text="运行日志", padding=10)
        frame_log.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.text_log = tk.Text(frame_log, height=14, state=tk.DISABLED, font=("Consolas", 10))
        self.text_log.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(self.text_log, command=self.text_log.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_log.config(yscrollcommand=scrollbar.set)

        self.selected_file = ""

    def log(self, msg):
        self.text_log.config(state=tk.NORMAL)
        self.text_log.insert(tk.END, msg + "\n")
        self.text_log.see(tk.END)
        self.text_log.config(state=tk.DISABLED)

    def select_file(self):
        path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if path:
            self.selected_file = path
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, path)
            self.log("已选择文件：%s" % path)

    def download_template(self):
        path = filedialog.asksaveasfilename(
            title="保存模板",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="用例需求溯源输入模板.xlsx"
        )
        if path:
            try:
                create_template(path)
                self.log("模板已下载：%s" % path)
                messagebox.showinfo("完成", "模板已保存至：\n%s" % path)
            except Exception as e:
                self.log("下载模板失败：%s" % e)
                messagebox.showerror("错误", str(e))

    def run_analysis(self):
        input_path = self.entry_path.get().strip()
        if not input_path or not os.path.exists(input_path):
            messagebox.showwarning("提示", "请先选择一个有效的Excel文件")
            return

        case_prefix = self.entry_case_prefix.get().strip()
        req_prefix = self.entry_req_prefix.get().strip()

        base, ext = os.path.splitext(input_path)
        output_path = "%s_output.xlsx" % base

        self.btn_run.config(state=tk.DISABLED)
        self.log("=" * 40)
        self.log("开始分析...")
        if case_prefix:
            self.log("用例ID前缀过滤：%s" % case_prefix)
        if req_prefix:
            self.log("需求ID前缀过滤：%s" % req_prefix)

        try:
            stats = analyze_traceability(input_path, output_path, case_prefix, req_prefix)
            self.log("分析完成")
            self.log("   输出文件：%s" % output_path)
            self.log("   有效用例数：%d" % stats['case_count'])
            self.log("   有效需求数：%d" % stats['req_count'])

            if stats['invalid_cases']:
                self.log("   用例ID异常（%d个）：%s" % (len(stats['invalid_cases']), ', '.join(stats['invalid_cases'])))

            if stats['invalid_reqs']:
                self.log("   需求ID异常（%d个）：" % len(stats['invalid_reqs']))
                for case_id, req_id in stats['invalid_reqs']:
                    self.log("      - 用例 [%s] 包含异常需求 [%s]" % (case_id, req_id))

            if stats['orphan_cases']:
                self.log("   孤儿用例（%d个，已移至异常项Sheet）：%s" % (len(stats['orphan_cases']), ', '.join(stats['orphan_cases'])))

            if stats['orphan_reqs']:
                self.log("   孤儿需求（%d个，已移至异常项Sheet）：%s" % (len(stats['orphan_reqs']), ', '.join(stats['orphan_reqs'])))

            if not stats['orphan_cases'] and not stats['orphan_reqs']:
                self.log("   未发现孤儿用例或孤儿需求")

            messagebox.showinfo("完成", "溯源分析已完成！\n输出文件：%s" % output_path)
        except Exception as e:
            self.log("错误：%s" % e)
            messagebox.showerror("错误", str(e))
        finally:
            self.btn_run.config(state=tk.NORMAL)


if __name__ == "__main__":
    root = tk.Tk()
    app = TraceabilityGUI(root)
    root.mainloop()
