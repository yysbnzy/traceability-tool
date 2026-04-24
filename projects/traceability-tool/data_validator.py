# -*- coding: utf-8 -*-
"""
数据校验模块 - 验证数据完整性
正确逻辑：双向溯源表中已包含所有用例（包括标记为孤儿用例的）
"""

from openpyxl import load_workbook
import re

def validate_data_integrity(output_file):
    """
    验证数据完整性
    """
    wb = load_workbook(output_file, data_only=True)
    
    sheet_names = wb.sheetnames
    print("Worksheet list:", [s for s in sheet_names])
    
    # 找到对应的工作表
    ws_bidirectional = None
    ws_exceptions = None
    ws_input = None
    
    for name in sheet_names:
        if '双向' in name:
            ws_bidirectional = wb[name]
        elif '异常' in name or '差异' in name:
            ws_exceptions = wb[name]
        elif '输入' in name:
            ws_input = wb[name]
    
    if not ws_bidirectional or not ws_input:
        print("Error: Required worksheets not found")
        return None
    
    # 统计双向溯源表（区分有需求和无需求的用例）
    cases_with_reqs = set()  # 有需求的用例
    orphan_cases_in_bidirectional = set()  # 孤儿用例（标记为孤儿用例）
    bidirectional_reqs = set()  # 所有需求
    
    for row in ws_bidirectional.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            case_id = str(row[0]).strip()
            req_text = str(row[1]) if len(row) > 1 and row[1] else ''
            
            if '(孤儿用例)' in req_text or req_text == '':
                orphan_cases_in_bidirectional.add(case_id)
            else:
                cases_with_reqs.add(case_id)
            
            # 统计需求（排除孤儿用例标记）
            if req_text and '(孤儿用例)' not in req_text:
                reqs = re.split(r'[,;\s]+', req_text)
                for req in reqs:
                    req = req.strip()
                    if req and req != '(无)':
                        bidirectional_reqs.add(req)
        
        # 需求->用例列
        if len(row) > 4 and row[3] and str(row[3]).strip():
            req_id = str(row[3]).strip()
            if '(孤儿需求)' not in req_id:
                bidirectional_reqs.add(req_id)
    
    # 统计异常分析表
    exception_orphan_cases = []
    exception_orphan_reqs = []
    if ws_exceptions:
        for row in ws_exceptions.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3:
                exc_type = str(row[0]) if row[0] else ''
                exc_id = str(row[1]) if row[1] else ''
                
                if '孤儿用例' in exc_type and exc_id:
                    exception_orphan_cases.append(exc_id)
                elif '孤儿需求' in exc_type and exc_id:
                    exception_orphan_reqs.append(exc_id)
    
    # 统计输入源表
    input_cases = set()
    input_reqs = set()
    input_orphan_cases = []
    
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2:
            case_id = str(row[0]) if row[0] else ''
            req_text = str(row[1]) if row[1] else ''
            
            if case_id and case_id != '(空)':
                input_cases.add(case_id)
                # 检查是否为孤儿用例
                if not req_text or req_text == '(无)' or req_text == '(空)':
                    input_orphan_cases.append(case_id)
            
            if req_text and req_text != '(无)' and req_text != '(空)':
                reqs = re.split(r'[,;\s]+', req_text)
                for req in reqs:
                    req = req.strip()
                    if req and req != '(无)':
                        input_reqs.add(req)
    
    # 验证
    print("\n" + "="*60)
    print("Data Integrity Validation Report")
    print("="*60)
    
    # 用例验证（正确逻辑：双向溯源表已包含所有用例）
    print(f"\n[Case Validation]")
    print(f"  Bidirectional table - Cases with requirements: {len(cases_with_reqs)}")
    print(f"  Bidirectional table - Orphan cases: {len(orphan_cases_in_bidirectional)}")
    print(f"  Bidirectional table - Total cases: {len(cases_with_reqs) + len(orphan_cases_in_bidirectional)}")
    print(f"  Exception table - Orphan cases: {len(exception_orphan_cases)}")
    print(f"  Input source - Total cases: {len(input_cases)}")
    print(f"  Input source - Orphan cases: {len(input_orphan_cases)}")
    
    # 验证1：双向溯源表总数 = 输入源总数
    total_bidirectional = len(cases_with_reqs) + len(orphan_cases_in_bidirectional)
    case_check_total = total_bidirectional == len(input_cases)
    
    # 验证2：双向溯源表孤儿用例 = 异常分析表孤儿用例 = 输入源孤儿用例
    orphan_check = (len(orphan_cases_in_bidirectional) == len(exception_orphan_cases) == len(input_orphan_cases))
    
    if case_check_total:
        print(f"  [PASS] Total cases match: {total_bidirectional} = {len(input_cases)}")
    else:
        print(f"  [FAIL] Total cases mismatch: {total_bidirectional} != {len(input_cases)}")
    
    if orphan_check:
        print(f"  [PASS] Orphan cases count match: {len(orphan_cases_in_bidirectional)}")
    else:
        print(f"  [WARN] Orphan cases count differ: Bidirectional={len(orphan_cases_in_bidirectional)}, Exception={len(exception_orphan_cases)}, Input={len(input_orphan_cases)}")
    
    # 需求验证
    print(f"\n[Requirement Validation]")
    print(f"  Bidirectional table requirements: {len(bidirectional_reqs)}")
    print(f"  Exception orphan requirements: {len(exception_orphan_reqs)}")
    print(f"  Total: {len(bidirectional_reqs) + len(exception_orphan_reqs)}")
    print(f"  Input source requirements: {len(input_reqs)}")
    
    req_check = len(bidirectional_reqs) + len(exception_orphan_reqs) == len(input_reqs)
    if req_check:
        print(f"  [PASS] Requirements match")
    else:
        print(f"  [FAIL] Requirements mismatch (diff: {abs(len(bidirectional_reqs) + len(exception_orphan_reqs) - len(input_reqs))})")
    
    print("\n" + "="*60)
    
    if case_check_total and req_check:
        print("[PASS] All data integrity checks passed!")
    else:
        print("[FAIL] Some data integrity checks failed!")
    
    print("="*60)
    
    return {
        'case_check_total': case_check_total,
        'orphan_check': orphan_check,
        'req_check': req_check,
        'cases_with_reqs': len(cases_with_reqs),
        'orphan_cases_bidirectional': len(orphan_cases_in_bidirectional),
        'orphan_cases_exception': len(exception_orphan_cases),
        'orphan_cases_input': len(input_orphan_cases),
        'input_cases': len(input_cases),
        'bidirectional_reqs': len(bidirectional_reqs),
        'exception_reqs': len(exception_orphan_reqs),
        'input_reqs': len(input_reqs)
    }

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        result = validate_data_integrity(sys.argv[1])
    else:
        result = validate_data_integrity(r'C:\Users\Administrator\.openclaw\media\temp\用例需求文档_溯源分析结果_最终版.xlsx')
