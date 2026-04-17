# -*- coding: utf-8 -*-
"""
测试用例与需求溯源工具 v3.0 - 支持跨文档列拼接
"""

import re
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


class TraceabilityTool:
    def __init__(self, root):
        self.root = root
        self.root.title("测试用例与需求溯源工具 v3.0")
        self.root.geometry("800x700")
        self.case_file_path = None
        self.req_file_path = None
        self.case_wb = None
        self.req_wb = None
        self.create_ui()
    
    def create_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(2, weight=1)
        row = 0
        
        # ===== 用例文档配置 =====
        ttk.Label(main_frame, text="【用例文档配置】", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        row += 1
        
        ttk.Label(main_frame, text="用例文档:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.case_file_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.case_file_var, width=50).grid(row=row, column=1, sticky=(tk.W, tk.E), pady=2, padx=5)
        ttk.Button(main_frame, text="浏览...", command=self.select_case_file).grid(row=row, column=2, pady=2)
        row += 1
        
        ttk.Label(main_frame, text="用例ID列:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.case_id_col = ttk.Combobox(main_frame, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=10, state="readonly")
        self.case_id_col.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.case_id_col.set("E")
        row += 1
        
        ttk.Label(main_frame, text="需求ID列:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.case_req_col = ttk.Combobox(main_frame, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=10, state="readonly")
        self.case_req_col.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.case_req_col.set("B")
        row += 1
        
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # ===== 用例ID拼接配置（新功能）=====
        ttk.Label(main_frame, text="【用例ID拼接配置】", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        row += 1
        
        # 用例前缀来源
        ttk.Label(main_frame, text="前缀来源:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.case_prefix_source = ttk.Combobox(main_frame, values=["固定文本", "用例文档列", "需求文档列"], width=15, state="readonly")
        self.case_prefix_source.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.case_prefix_source.set("固定文本")
        self.case_prefix_source.bind("<<ComboboxSelected>>", lambda e: self.on_source_change("case_prefix"))
        
        self.case_prefix_value_frame = ttk.Frame(main_frame)
        self.case_prefix_value_frame.grid(row=row, column=2, sticky=tk.W, pady=2)
        self.case_prefix_text = ttk.Entry(self.case_prefix_value_frame, width=20)
        self.case_prefix_text.pack(side=tk.LEFT)
        row += 1
        
        # 用例后缀来源
        ttk.Label(main_frame, text="后缀来源:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.case_suffix_source = ttk.Combobox(main_frame, values=["固定文本", "用例文档列", "需求文档列"], width=15, state="readonly")
        self.case_suffix_source.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.case_suffix_source.set("固定文本")
        self.case_suffix_source.bind("<<ComboboxSelected>>", lambda e: self.on_source_change("case_suffix"))
        
        self.case_suffix_value_frame = ttk.Frame(main_frame)
        self.case_suffix_value_frame.grid(row=row, column=2, sticky=tk.W, pady=2)
        self.case_suffix_text = ttk.Entry(self.case_suffix_value_frame, width=20)
        self.case_suffix_text.pack(side=tk.LEFT)
        row += 1
        
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # ===== 需求ID拼接配置（新功能）=====
        ttk.Label(main_frame, text="【需求ID拼接配置】", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        row += 1
        
        # 需求前缀来源
        ttk.Label(main_frame, text="前缀来源:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.req_prefix_source = ttk.Combobox(main_frame, values=["固定文本", "用例文档列", "需求文档列"], width=15, state="readonly")
        self.req_prefix_source.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.req_prefix_source.set("固定文本")
        self.req_prefix_source.bind("<<ComboboxSelected>>", lambda e: self.on_source_change("req_prefix"))
        
        self.req_prefix_value_frame = ttk.Frame(main_frame)
        self.req_prefix_value_frame.grid(row=row, column=2, sticky=tk.W, pady=2)
        self.req_prefix_text = ttk.Entry(self.req_prefix_value_frame, width=20)
        self.req_prefix_text.pack(side=tk.LEFT)
        row += 1
        
        # 需求后缀来源
        ttk.Label(main_frame, text="后缀来源:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.req_suffix_source = ttk.Combobox(main_frame, values=["固定文本", "用例文档列", "需求文档列"], width=15, state="readonly")
        self.req_suffix_source.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.req_suffix_source.set("固定文本")
        self.req_suffix_source.bind("<<ComboboxSelected>>", lambda e: self.on_source_change("req_suffix"))
        
        self.req_suffix_value_frame = ttk.Frame(main_frame)
        self.req_suffix_value_frame.grid(row=row, column=2, sticky=tk.W, pady=2)
        self.req_suffix_text = ttk.Entry(self.req_suffix_value_frame, width=20)
        self.req_suffix_text.pack(side=tk.LEFT)
        row += 1
        
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # ===== 需求文档配置（可选）=====
        ttk.Label(main_frame, text="【需求文档配置（可选）】", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        row += 1
        
        ttk.Label(main_frame, text="需求文档:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.req_file_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.req_file_var, width=50).grid(row=row, column=1, sticky=(tk.W, tk.E), pady=2, padx=5)
        ttk.Button(main_frame, text="浏览...", command=self.select_req_file).grid(row=row, column=2, pady=2)
        row += 1
        
        ttk.Label(main_frame, text="需求ID列:").grid(row=row, column=0, sticky=tk.W, pady=2)
        self.req_id_col = ttk.Combobox(main_frame, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=10, state="readonly")
        self.req_id_col.grid(row=row, column=1, sticky=tk.W, pady=2, padx=5)
        self.req_id_col.set("A")
        row += 1
        
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        # ===== 按钮 =====
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=3, pady=20)
        ttk.Button(button_frame, text="开始溯源分析", command=self.start_analysis, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="退出", command=self.root.quit, width=10).pack(side=tk.LEFT, padx=5)
    
    def on_source_change(self, field):
        """当来源类型改变时，切换输入控件"""
        source_map = {
            "case_prefix": (self.case_prefix_source, self.case_prefix_value_frame, self.case_prefix_text),
            "case_suffix": (self.case_suffix_source, self.case_suffix_value_frame, self.case_suffix_text),
            "req_prefix": (self.req_prefix_source, self.req_prefix_value_frame, self.req_prefix_text),
            "req_suffix": (self.req_suffix_source, self.req_suffix_value_frame, self.req_suffix_text),
        }
        
        source_combo, frame, text_entry = source_map[field]
        source_type = source_combo.get()
        
        # 清除旧控件
        for widget in frame.winfo_children():
            widget.destroy()
        
        if source_type == "固定文本":
            # 文本输入框
            new_entry = ttk.Entry(frame, width=20)
            new_entry.pack(side=tk.LEFT)
            # 更新引用
            if field == "case_prefix":
                self.case_prefix_text = new_entry
            elif field == "case_suffix":
                self.case_suffix_text = new_entry
            elif field == "req_prefix":
                self.req_prefix_text = new_entry
            elif field == "req_suffix":
                self.req_suffix_text = new_entry
        else:
            # 列选择下拉框
            col_combo = ttk.Combobox(frame, values=list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"), width=10, state="readonly")
            col_combo.pack(side=tk.LEFT)
            col_combo.set("A")
            # 更新引用
            if field == "case_prefix":
                self.case_prefix_text = col_combo
            elif field == "case_suffix":
                self.case_suffix_text = col_combo
            elif field == "req_prefix":
                self.req_prefix_text = col_combo
            elif field == "req_suffix":
                self.req_suffix_text = col_combo
    
    def select_case_file(self):
        file_path = filedialog.askopenfilename(title="选择用例文档", filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if file_path:
            self.case_file_path = file_path
            self.case_file_var.set(file_path)
            try:
                self.case_wb = load_workbook(file_path, read_only=True, data_only=True)
            except Exception as e:
                messagebox.showerror("错误", f"加载用例文档失败: {str(e)}")
    
    def select_req_file(self):
        file_path = filedialog.askopenfilename(title="选择需求文档", filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if file_path:
            self.req_file_path = file_path
            self.req_file_var.set(file_path)
            try:
                self.req_wb = load_workbook(file_path, read_only=True, data_only=True)
            except Exception as e:
                messagebox.showerror("错误", f"加载需求文档失败: {str(e)}")
    
    def split_requirement_ids(self, text):
        if text is None:
            return []
        normalized = str(text)
        normalized = re.sub(r'[，；。\s\n]+', ',', normalized)
        parts = re.split(r'[,;]+', normalized)
        ids = [p.strip() for p in parts if p.strip()]
        return ids
    
    def col_to_index(self, col):
        return ord(col) - ord('A')
    
    def get_value_from_source(self, source_type, value_widget, row_data_case, row_data_req):
        """根据来源类型获取值"""
        if source_type == "固定文本":
            return value_widget.get().strip()
        elif source_type == "用例文档列":
            col_idx = self.col_to_index(value_widget.get())
            if row_data_case and len(row_data_case) > col_idx:
                val = row_data_case[col_idx]
                return str(val).strip() if val else ""
            return ""
        elif source_type == "需求文档列":
            if not self.req_wb:
                return None  # 标记需要提示
            col_idx = self.col_to_index(value_widget.get())
            if row_data_req and len(row_data_req) > col_idx:
                val = row_data_req[col_idx]
                return str(val).strip() if val else ""
            return ""
        return ""
    
    def start_analysis(self):
        if not self.case_file_path:
            messagebox.showerror("错误", "请选择用例文档")
            return
        
        # 检查是否需要需求文档但未上传
        sources = [
            self.case_prefix_source.get(),
            self.case_suffix_source.get(),
            self.req_prefix_source.get(),
            self.req_suffix_source.get()
        ]
        if "需求文档列" in sources and not self.req_wb:
            messagebox.showwarning("提示", "您选择了'需求文档列'作为来源，但未上传需求文档。\n请先上传需求文档，或修改来源为其他选项。")
            return
        
        case_id_idx = self.col_to_index(self.case_id_col.get())
        case_req_idx = self.col_to_index(self.case_req_col.get())
        req_id_idx = self.col_to_index(self.req_id_col.get())
        
        try:
            self.analyze(case_id_idx, case_req_idx, req_id_idx)
            messagebox.showinfo("完成", "溯源分析已完成！")
        except Exception as e:
            messagebox.showerror("错误", f"分析失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def analyze(self, case_id_idx, case_req_idx, req_id_idx):
        case_to_reqs = {}
        case_to_reqs_raw = {}
        req_to_cases = {}
        
        case_ws = self.case_wb.active
        max_col = max(case_id_idx, case_req_idx) + 1
        
        # 获取来源配置
        case_prefix_source = self.case_prefix_source.get()
        case_suffix_source = self.case_suffix_source.get()
        req_prefix_source = self.req_prefix_source.get()
        req_suffix_source = self.req_suffix_source.get()
        
        # 读取需求文档数据（如果配置了需求文档列）
        req_ws = None
        if self.req_wb and ("需求文档列" in [case_prefix_source, case_suffix_source, req_prefix_source, req_suffix_source]):
            req_ws = self.req_wb.active
        
        # 为每行用例数据匹配需求文档数据（简单按行号匹配）
        req_data_cache = {}
        if req_ws:
            for idx, row in enumerate(req_ws.iter_rows(min_row=2, values_only=True), start=2):
                req_data_cache[idx] = row
        
        for idx, row in enumerate(case_ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
            if len(row) <= max(case_id_idx, case_req_idx):
                continue
            
            case_id_raw = str(row[case_id_idx]).strip() if row[case_id_idx] else ''
            req_text = row[case_req_idx]
            req_ids = self.split_requirement_ids(req_text)
            
            # 跳过空行
            if (not case_id_raw or case_id_raw == 'None') and not req_ids:
                continue
            
            # 获取当前行对应的需求文档行数据
            req_row_data = req_data_cache.get(idx, None)
            
            # 动态获取用例前缀和后缀
            case_prefix = self.get_value_from_source(case_prefix_source, self.case_prefix_text, row, req_row_data)
            case_suffix = self.get_value_from_source(case_suffix_source, self.case_suffix_text, row, req_row_data)
            
            # 处理用例ID
            case_id_full = None
            if case_id_raw and case_id_raw != 'None':
                case_id_full = f"{case_prefix}{case_id_raw}{case_suffix}"
                if case_id_full not in case_to_reqs:
                    case_to_reqs[case_id_full] = set()
                    case_to_reqs_raw[case_id_raw] = set()
            
            # 处理需求ID（每个需求独立获取前缀后缀）
            for req_id in req_ids:
                req_prefix = self.get_value_from_source(req_prefix_source, self.req_prefix_text, row, req_row_data)
                req_suffix = self.get_value_from_source(req_suffix_source, self.req_suffix_text, row, req_row_data)
                req_id_full = f"{req_prefix}{req_id}{req_suffix}"
                
                if req_id_full not in req_to_cases:
                    req_to_cases[req_id_full] = set()
                if case_id_full:
                    req_to_cases[req_id_full].add(case_id_full)
                
                if case_id_full:
                    case_to_reqs[case_id_full].add(req_id_full)
                    case_to_reqs_raw[case_id_raw].add(req_id_full)
        
        # 读取需求文档所有需求ID（用于计算缺失）
        all_req_ids_raw = set()
        if self.req_wb:
            req_ws = self.req_wb.active
            for row in req_ws.iter_rows(min_row=2, max_col=req_id_idx+1, values_only=True):
                if len(row) <= req_id_idx:
                    continue
                req_id = str(row[req_id_idx]).strip() if row[req_id_idx] else ''
                if req_id and req_id != 'None':
                    all_req_ids_raw.add(req_id)
        
        missing_case_reqs = all_req_ids_raw - set(req_to_cases.keys())
        
        output_path = filedialog.asksaveasfilename(title="保存分析结果", defaultextension=".xlsx", filetypes=[("Excel文件", "*.xlsx")])
        if not output_path:
            return
        
        self.export(output_path, case_to_reqs, case_to_reqs_raw, req_to_cases, all_req_ids_raw, missing_case_reqs)
    
    def export(self, output_path, case_to_reqs, case_to_reqs_raw, req_to_cases, all_req_ids_raw, missing_case_reqs):
        wb = Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 双向溯源表
        ws1 = wb.active
        ws1.title = "双向溯源表"
        ws1['A1'] = '用例ID'
        ws1['B1'] = '关联需求ID'
        for cell in [ws1['A1'], ws1['B1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        sorted_cases = sorted([case_id for case_id in case_to_reqs.keys() if case_to_reqs[case_id]])
        row_idx = 2
        for case_id in sorted_cases:
            reqs = sorted(case_to_reqs[case_id])
            ws1.cell(row=row_idx, column=1, value=case_id)
            ws1.cell(row=row_idx, column=2, value=", ".join(reqs))
            ws1.cell(row=row_idx, column=1).border = thin_border
            ws1.cell(row=row_idx, column=2).border = thin_border
            row_idx += 1
        
        ws1['D1'] = '需求ID'
        ws1['E1'] = '关联用例ID'
        for cell in [ws1['D1'], ws1['E1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        sorted_reqs = sorted([req_id for req_id in req_to_cases.keys() if req_to_cases[req_id]])
        for idx, req_id in enumerate(sorted_reqs, start=2):
            cases = sorted(req_to_cases[req_id])
            ws1.cell(row=idx, column=4, value=req_id)
            ws1.cell(row=idx, column=5, value=", ".join(cases))
            ws1.cell(row=idx, column=4).border = thin_border
            ws1.cell(row=idx, column=5).border = thin_border
        
        for col in ['A', 'B', 'D', 'E']:
            ws1.column_dimensions[col].width = 40
        
        # 输入源
        ws2 = wb.create_sheet("输入源")
        ws2['A1'] = '用例ID'
        ws2['B1'] = '需求ID'
        for cell in [ws2['A1'], ws2['B1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        case_ws = self.case_wb.active
        max_col = max(self.col_to_index(self.case_id_col.get()), self.col_to_index(self.case_req_col.get())) + 1
        for idx, row in enumerate(case_ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
            case_id_idx = self.col_to_index(self.case_id_col.get())
            case_req_idx = self.col_to_index(self.case_req_col.get())
            case_val = str(row[case_id_idx]) if len(row) > case_id_idx and row[case_id_idx] else ''
            req_val = str(row[case_req_idx]) if len(row) > case_req_idx and row[case_req_idx] else ''
            ws2.cell(row=idx, column=1, value=case_val)
            ws2.cell(row=idx, column=2, value=req_val)
            ws2.cell(row=idx, column=1).border = thin_border
            ws2.cell(row=idx, column=2).border = thin_border
        
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 35
        
        # 异常分析
        ws3 = wb.create_sheet("异常分析")
        ws3['A1'] = '异常类型'
        ws3['B1'] = 'ID'
        ws3['C1'] = '说明'
        for cell in [ws3['A1'], ws3['B1'], ws3['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row_idx = 2
        orphan_cases_raw = [c for c, r in case_to_reqs_raw.items() if not r]
        orphan_reqs = [r for r, c in req_to_cases.items() if not c]
        
        for case_id_raw in orphan_cases_raw:
            ws3.cell(row=row_idx, column=1, value="孤儿用例")
            ws3.cell(row=row_idx, column=2, value=case_id_raw)
            ws3.cell(row=row_idx, column=3, value="该用例未关联任何需求")
            for c in [1, 2, 3]:
                ws3.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
        
        for req_id in orphan_reqs:
            ws3.cell(row=row_idx, column=1, value="孤儿需求")
            ws3.cell(row=row_idx, column=2, value=req_id)
            ws3.cell(row=row_idx, column=3, value="该需求在用例文档中被引用，但无实际关联用例")
            for c in [1, 2, 3]:
                ws3.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
        
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 40
        ws3.column_dimensions['C'].width = 40
        
        # 需求缺失用例
        ws4 = wb.create_sheet("需求缺失用例")
        ws4['A1'] = '需求ID'
        ws4['B1'] = '状态'
        ws4['C1'] = '说明'
        for cell in [ws4['A1'], ws4['B1'], ws4['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row_idx = 2
        for req_id in sorted(missing_case_reqs):
            ws4.cell(row=row_idx, column=1, value=req_id)
            ws4.cell(row=row_idx, column=2, value="未覆盖")
            ws4.cell(row=row_idx, column=3, value="该需求在需求文档中存在，但未在测试用例中体现")
            for c in [1, 2, 3]:
                ws4.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
        
        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 50
        
        # 统计汇总
        ws5 = wb.create_sheet("统计汇总")
        ws5['A1'] = '统计项'
        ws5['B1'] = '数量'
        ws5['C1'] = '说明'
        for cell in [ws5['A1'], ws5['B1'], ws5['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        total_cases = len(case_to_reqs_raw)
        total_reqs_in_case = len(req_to_cases)
        linked_cases = len([c for c, r in case_to_reqs_raw.items() if r])
        linked_reqs = len([r for r, c in req_to_cases.items() if c])
        orphan_case_count = len(orphan_cases_raw)
        orphan_req_count = len(orphan_reqs)
        missing_req_count = len(missing_case_reqs)
        coverage_rate = (linked_reqs / len(all_req_ids_raw) * 100) if all_req_ids_raw else 0
        
        stats = [
            ('总用例数', total_cases, '用例文档中用例ID非空的总数'),
            ('双向关联用例数', linked_cases, '有关联需求的用例数'),
            ('孤儿用例数', orphan_case_count, '未关联任何需求的用例数'),
            ('', '', ''),
            ('用例中引用需求数', total_reqs_in_case, '用例文档中引用的需求ID总数'),
            ('双向关联需求数', linked_reqs, '有关联用例的需求数'),
            ('孤儿需求数', orphan_req_count, '有用例ID为空但引用的需求数'),
            ('', '', ''),
            ('需求文档总需求数', len(all_req_ids_raw), '需求文档中的需求总数（需上传需求文档）'),
            ('需求缺失用例数', missing_req_count, '需求文档中有但用例未覆盖的需求数'),
            ('需求覆盖率', f'{coverage_rate:.1f}%', '已覆盖需求数/需求文档总需求数'),
        ]
        
        row_idx = 2
        for item, count, desc in stats:
            ws5.cell(row=row_idx, column=1, value=item)
            ws5.cell(row=row_idx, column=2, value=count)
            ws5.cell(row=row_idx, column=3, value=desc)
            for c in [1, 2, 3]:
                ws5.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
        
        ws5.column_dimensions['A'].width = 25
        ws5.column_dimensions['B'].width = 15
        ws5.column_dimensions['C'].width = 50
        
        wb.save(output_path)
        print(f"溯源分析完成，结果已保存至：{output_path}")


def main():
    root = tk.Tk()
    app = TraceabilityTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
